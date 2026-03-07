"""
parse_eval.py — Generate car_eval.csv and pen_eval.csv from source Excel files.

Usage (from repo root):
    python eval/parse_eval.py

Inputs (repo root):
    so101_car_pick_and_placeeval.xlsx
    so101_pick_peneval.xlsx
    Trained_Models.xlsx

Outputs:
    eval/car_eval.csv
    eval/pen_eval.csv
"""

import csv, os, re, datetime
import openpyxl

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EVAL_DIR  = os.path.dirname(os.path.abspath(__file__))

CAR_XLSX = os.path.join(REPO_ROOT, 'so101_car_pick_and_placeeval.xlsx')
PEN_XLSX = os.path.join(REPO_ROOT, 'so101_pick_peneval.xlsx')

HF = 'jonathm126'

# ── Only sheets/models that appear in Experiment Plan col G ──────────────────
# Maps eval-sheet Version cell value -> (full_hf_model_id, hf_eval_dataset)

CAR_MODEL_META = {
    '50_episodes_v2': (
        f'{HF}/act-so101_car_pick_and_place-50_episodes_v2',
        f'{HF}/eval_so101_car_pick_and_place_50_episodes_v2_12_test_case'),
    '50_episodes_v4': (
        f'{HF}/act-so101_car_pick_and_place-50_episodes_v4',
        f'{HF}/eval_so101_car_pick_and_place_50_episodes_v4_real_v0'),
    '25_episodes_v0': (
        f'{HF}/act-so101_car_pick_and_place-25_episodes_v0',
        f'{HF}/eval_so101_car_pick_and_place_25_episodes_v0_real_v0'),
    '96_episodes_v0': (
        f'{HF}/act-so101_car_pick_and_place-96_episodes_v0',
        f'{HF}/eval_so101_car_pick_and_place-96_episodes_v0-real_v0'),
    '96_episodes_v1': (
        f'{HF}/act-so101_car_pick_and_place-96_episodes_v1',
        f'{HF}/eval_so101_car_pick_and_place_96_episodes_v1_real_v0'),
    'yolo_v3': (
        f'{HF}/act-so101_car_pick_and_place-bbox-yolo_v3',
        f'{HF}/eval_so101_car_pick_and_place-bbox_yolo_v3-real_v0'),
}

PEN_MODEL_META = {
    'baseline_v0': (
        f'{HF}/act-so101_pick_pen-baseline_v0',
        f'{HF}/eval_so101_pick_pen-v0-real_v0'),
    'v1': (
        f'{HF}/act-so101_pick_pen-v1',
        f'{HF}/eval_so101_pick_pen_v1_real_v0'),
    'pick_pen-v3': (
        f'{HF}/act-so101_pick_pen-v3',
        f'{HF}/eval_so101_pick_pen_v3_real_v0'),
    'pick_pen-v4': (
        f'{HF}/act-so101_pick_pen-v4',
        f'{HF}/eval_so101_pick_pen_v4_real_v0'),
    'pick_pen-v5': (
        f'{HF}/act-so101_pick_pen-v5',
        f'{HF}/eval_so101_pick_pen_v5_real_v0'),
    'yolo_v0': (
        f'{HF}/act-so101_pick_pen-bbox-yolo_v0',
        f'{HF}/eval_so101_pick_pen-bbox_yolo_v0-real_v0'),
    '25_episodes_v0': (
        f'{HF}/act-so101_pick_pen-25_episodes_v0',
        f'{HF}/eval_so101_pick_pen_25_episodes_v0_real_v0'),
}

# ── Checkpoint normalisation: anything like "60k","60K" -> 60000 ──────────────
def norm_ckpt(val):
    s = str(val).strip().upper()
    m = re.fullmatch(r'(\d+(?:\.\d+)?)K', s)
    if m:
        return str(int(float(m.group(1)) * 1000))
    return s

# ── Pose normalisation ────────────────────────────────────────────────────────
MONTH = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
         'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

def clean_pose(val):
    """Normalise pose values: Excel date misparses -> 'x-y', formulas -> number."""
    if val is None: return ''
    if isinstance(val, datetime.datetime):
        return f'{val.month}-{val.day}'
    s = str(val).strip()
    # Excel formula like '=45+90' or '=90+45'
    m = re.fullmatch(r'=(\d+)\+(\d+)', s)
    if m: return str(int(m.group(1)) + int(m.group(2)))
    # Date-like strings: '2-Jan', 'Jan-2'
    m = re.fullmatch(r'(\d+)-([a-zA-Z]+)', s)
    if m:
        mo = MONTH.get(m.group(2).lower())
        if mo: return f'{mo}-{m.group(1)}'
    m = re.fullmatch(r'([a-zA-Z]+)-(\d+)', s)
    if m:
        mo = MONTH.get(m.group(1).lower())
        if mo: return f'{mo}-{m.group(2)}'
    return s

def is_oos(pose_val, is_rotation=False):
    s = clean_pose(pose_val)
    if '-' in s or '=' in s: return True
    try:
        v = int(float(s))
        if is_rotation: return v not in (0, 90, 180, 270)
    except (ValueError, TypeError):
        return True
    return False

# ── Column accessor ───────────────────────────────────────────────────────────
CAR_ALIASES = {'v':'Version','Bin':'Target Position','Start':'Source Position',
               'Rotation':'Source Orientation','Score_v1':'Score','Score_v2':'Score',
               'score_v0':None,'episode':'Episode'}

def norm_header(h, aliases): return [aliases.get(c,c) if c else None for c in h]

def get(row, header, name):
    try: idx = header.index(name); return row[idx] if idx < len(row) else None
    except ValueError: return None

def safe(v): return '' if v is None else str(v)

# ── CAR ───────────────────────────────────────────────────────────────────────
HEADER_CAR = ['task','hf_model_id','hf_eval_dataset','checkpoint','K','TE',
              'episode','target_pos','source_pos','orientation',
              'score','eval_type','experiment_note','row_note']

wb = openpyxl.load_workbook(CAR_XLSX)
car_rows = []

for sheet in wb.sheetnames:
    ws = wb[sheet]
    raw = [r for r in ws.iter_rows(min_row=1, values_only=True) if any(v is not None for v in r)]
    if not raw: continue
    header = norm_header(raw[0], CAR_ALIASES)

    for row in raw[1:]:
        version = safe(get(row, header, 'Version'))
        if not version or version in ('Version','v'): continue

        # Skip models not in experiment plan
        if version not in CAR_MODEL_META: continue

        score_raw = get(row, header, 'Score')
        # Skip rows with no score
        if not isinstance(score_raw, (int, float)): continue

        hf_model, hf_eval = CAR_MODEL_META[version]
        ckpt   = norm_ckpt(safe(get(row, header, 'Checkpoint')))
        K      = safe(get(row, header, 'n_action_steps'))
        TE     = safe(get(row, header, 'temporal_ensembling'))
        target = clean_pose(get(row, header, 'Target Position'))
        source = clean_pose(get(row, header, 'Source Position'))
        orient = clean_pose(get(row, header, 'Source Orientation'))
        ep     = safe(get(row, header, 'Episode'))
        score  = str(score_raw)

        oos = is_oos(source) or is_oos(orient, is_rotation=True)
        eval_type = 'out_of_sample' if oos else 'in_sample'

        # experiment note only for sheets with multiple configs
        exp_note = ''
        if sheet == '96_episodes_v0':
            exp_note = f'K={K}, TE={TE}, ckpt={ckpt}'
        elif sheet == '50_episodes_v2' and K == '50':
            exp_note = 'K=50 ablation'

        extra = [safe(row[i]) for i in range(9, len(row))
                 if i < len(row) and row[i] is not None
                 and str(row[i]).strip() not in ('','None')]
        row_note = '; '.join(extra)

        car_rows.append([
            'car_pick_and_place', hf_model, hf_eval,
            ckpt, K, TE, ep, target, source, orient,
            score, eval_type, exp_note, row_note
        ])

with open(os.path.join(EVAL_DIR, 'car_eval.csv'), 'w', newline='') as f:
    csv.writer(f).writerows([HEADER_CAR] + car_rows)
print(f'car_eval.csv: {len(car_rows)} rows')

# ── PEN ───────────────────────────────────────────────────────────────────────
HEADER_PEN = ['task','hf_model_id','hf_eval_dataset','checkpoint','K','TE',
              'episode','position','rotation',
              'score','eval_type','experiment_note','row_note']

wb2 = openpyxl.load_workbook(PEN_XLSX)
pen_rows = []

for sheet in wb2.sheetnames:
    ws = wb2[sheet]
    raw = [r for r in ws.iter_rows(min_row=1, values_only=True) if any(v is not None for v in r)]
    if not raw: continue
    header = list(raw[0])

    for row in raw[1:]:
        version = safe(get(row, header, 'Version'))
        if not version or version == 'Version': continue

        # Skip models not in experiment plan
        if version not in PEN_MODEL_META: continue

        score_raw = get(row, header, 'Score')
        if not isinstance(score_raw, (int, float)): continue

        hf_model, hf_eval = PEN_MODEL_META[version]
        raw_ckpt = get(row, header, 'Checkpoint')
        if raw_ckpt is None: continue   # corrupt row (e.g. v2 sheet)
        ckpt  = norm_ckpt(safe(raw_ckpt))
        K     = safe(get(row, header, 'n_action_steps'))
        TE    = safe(get(row, header, 'temporal_ensembling'))
        pos   = clean_pose(get(row, header, 'Position'))
        rot   = clean_pose(get(row, header, 'Rotation'))
        ep    = safe(get(row, header, 'episode'))
        score = str(score_raw)

        oos = is_oos(pos) or is_oos(rot, is_rotation=True)
        eval_type = 'out_of_sample' if oos else 'in_sample'

        exp_note = ''
        if sheet == 'v4':
            exp_note = f'K={K}, TE={TE}'

        extra = [safe(row[i]) for i in range(8, len(row))
                 if i < len(row) and row[i] is not None
                 and '=AVERAGE' not in str(row[i])
                 and str(row[i]).strip() not in ('','None')]
        row_note = '; '.join(extra)

        pen_rows.append([
            'pick_pen', hf_model, hf_eval,
            ckpt, K, TE, ep, pos, rot,
            score, eval_type, exp_note, row_note
        ])

with open(os.path.join(EVAL_DIR, 'pen_eval.csv'), 'w', newline='') as f:
    csv.writer(f).writerows([HEADER_PEN] + pen_rows)
print(f'pen_eval.csv: {len(pen_rows)} rows')
